from keras.preprocessing.image import ImageDataGenerator
from keras.layers import Conv2D, MaxPooling2D, Dropout, Dense, Flatten
from keras.models import Sequential
from keras import backend as K
import os
from openpyxl import Workbook
import matplotlib.pyplot as plt

plt.rcParams["axes.labelsize"] = 16
plt.rcParams["xtick.labelsize"] = 14
plt.rcParams["ytick.labelsize"] = 14

K.clear_session()

trainingData = "./cnn/data/training"  # It contains 9081 images
validationData = "./cnn/data/validation"  # It contains 3632 images
directory = "./cnn/model/"

trainDataGen = ImageDataGenerator(
    rescale=1.0 / 255,
    rotation_range=12,
    width_shift_range=0.2,
    height_shift_range=0.2,
    zoom_range=0.15,
    horizontal_flip=True,
)

validationDataGen = ImageDataGenerator(rescale=1.0 / 255)

trainGenerated = trainDataGen.flow_from_directory(
    trainingData,
    target_size=(256, 256),
    batch_size=32,
    color_mode="grayscale",
    class_mode="categorical",
)
validationGenerated = validationDataGen.flow_from_directory(
    validationData,
    target_size=(256, 256),
    batch_size=32,
    color_mode="grayscale",
    class_mode="categorical",
)

print(validationGenerated.class_indices)  # labels from classes
print(trainGenerated.class_indices)

for imageBatch, labelsBatch in trainGenerated:
    print(imageBatch.shape)
    print(labelsBatch.shape)
    break

trainSteps = trainGenerated.n // trainGenerated.batch_size
validationSteps = validationGenerated.n // validationGenerated.batch_size


model = Sequential()
model.add(Conv2D(32, (5, 5), activation="relu", input_shape=(256, 256, 1)))
model.add(MaxPooling2D((2, 2)))
model.add(Conv2D(64, (3, 3), activation="relu"))
model.add(Conv2D(64, (3, 3), activation="relu"))
model.add(MaxPooling2D((2, 2)))
model.add(Conv2D(128, (3, 3), activation="relu"))
model.add(MaxPooling2D((2, 2)))
model.add(Conv2D(256, (3, 3), activation="relu"))
model.add(MaxPooling2D((2, 2)))
model.add(Flatten())
model.add(Dense(128, activation="relu"))
model.add(Dropout(0.25))
model.add(Dense(6, activation="softmax"))
model.summary()

model.compile(optimizer="adam", loss="categorical_crossentropy", metrics=["acc"])

history = model.fit(
    trainGenerated,
    steps_per_epoch=trainSteps,
    epochs=10,
    validation_data=validationGenerated,
    validation_steps=validationSteps,
)

if not os.path.exists(directory):
    os.mkdir(directory)
model.save("./cnn/model/model.h5", save_format="h5")
model.save_weights("./cnn/model/weights.h5", save_format="h5")

historyDictionary = history.history
print(historyDictionary.keys())

acc = history.history["acc"]
valAccuracy = history.history["val_acc"]  # final accuracy!
loss = history.history["loss"]
valLoss = history.history["val_loss"]

print(acc)
print(valAccuracy)
print(loss)
print(valLoss)

print("----------------- MODEL RESULT -----------------")
print()
print()
print()
print()
print("The final model accuracy is: ", historyDictionary["val_acc"][-1])

plt.figure(figsize=(16, 6))
plt.subplot(1, 2, 1)
epochs = len(history.history["loss"])
plt.plot(range(epochs), history.history["loss"], "g-", label="training loss")
plt.plot(range(epochs), history.history["val_loss"], "c-", label="validation loss")
plt.legend(prop={"size": 20})
plt.title("Training and validation loss")
plt.ylabel("loss")
plt.xlabel("number of epochs")
plt.subplot(1, 2, 2)
plt.plot(range(epochs), history.history["acc"], "g-", label="training accuracy")
plt.plot(range(epochs), history.history["val_acc"], "c-", label="validation accuracy")
plt.legend(prop={"size": 20})
plt.title("Training and validation accuracy")
plt.ylabel("accuracy")
plt.xlabel("number of epochs")
plt.savefig("./cnn/model/results_graph.png")
plt.show()

wb = Workbook()

hoja = wb.active
hoja.title = "Training Record"

num = 0
fila = 2
accuracy = 2
val_ac = 3
val_l = 4
lss = 5

for accu, vala, vals, loss in zip(acc, valAccuracy, valLoss, loss):
    hoja.cell(row=1, column=2, value="Accuracy")
    hoja.cell(row=1, column=3, value="Val_accuracy")
    hoja.cell(row=1, column=4, value="Val_loss")
    hoja.cell(row=1, column=5, value="Loss")

    hoja.cell(column=1, row=fila, value=num)
    hoja.cell(column=accuracy, row=fila, value=accu)
    hoja.cell(column=val_ac, row=fila, value=vala)
    hoja.cell(column=val_l, row=fila, value=vals)
    hoja.cell(column=lss, row=fila, value=loss)
    fila += 1
    num += 1

wb.save("./cnn/model/results.xlsx")